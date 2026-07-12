import json, sys
from copy import copy
from openpyxl import load_workbook
template, output, raw = sys.argv[1], sys.argv[2], sys.argv[3]
values = json.loads(raw)
wb = load_workbook(template); ws = wb.active
header = next((r for r in range(1, min(ws.max_row,20)+1) if any(ws.cell(r,c).value for c in range(1,ws.max_column+1))),1)
columns = {str(ws.cell(header,c).value).strip():c for c in range(1,ws.max_column+1) if ws.cell(header,c).value is not None}
missing = [x for x in values if x not in columns]
if missing: raise SystemExit('Template is missing columns: '+', '.join(missing))
row = max(ws.max_row+1,header+1)
for c in range(1,ws.max_column+1): ws.cell(row,c)._style=copy(ws.cell(row-1,c)._style)
for key,value in values.items(): ws.cell(row,columns[key]).value=value
wb.save(output)
